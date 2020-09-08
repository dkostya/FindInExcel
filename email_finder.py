# coding=utf-8
import openpyxl
import pkg_resources.py2_warn
from progress.bar import IncrementalBar

print('Для обработки необходимо поместить файл с базой и файл со списком e-mail')
print('в директорию, где находится файл программы обработчика email_finder.exe.')
print()
print('Файлы с паролем на открытие не поддерживаются!!')
print()
base_name = input('Введите имя файла с базой клиентов: ')
emails_name = input('Введите имя файла со списком e-mail: ')

print('Загрузка базы, подождите...')
base = openpyxl.load_workbook(base_name)

emails = (openpyxl.load_workbook(emails_name)).active

results = openpyxl.Workbook()
results_sheet = results.active

results_sheet.title = 'Results'
results_sheet.append(['Имя листа', 'e-mail', 'Дата обновления'])

print('Начинаю обработку')

for sheet in base.worksheets: # поиск email в базе
    print('Лист:', sheet.title, "Строк:", sheet.max_row, "Столбцов:", sheet.max_column)

    bar = IncrementalBar(' Progess', max=sheet.max_row)

    for row in sheet.iter_rows(max_col=50, values_only=True):

        bar.next()
        date = row[0]
        for cell in row:

            for col in emails.iter_cols(max_col=1, values_only=True):  # список email
                for email in col:
                    if cell == email:
                        new_row = []
                        new_row.append(sheet.title)
                        new_row.append(email)
                        new_row.append(date)
                        results_sheet.append(new_row)

    bar.finish()

print('Завершено')
print("Результаты отбработки сохранены в файле results.xlsx")
input('Для выхода нажмите любую кнопку')

results.save('results.xlsx')


